from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from .models import RawAttendance, Attendance, DailyRecord
from employee.models import Employee
from employee.filters import EmployeeFilter
from device.models import Device
from overtime.models import Overtime
from shift.models import Shift
from leave.models import Leave
from holiday.models import Holiday
from django.core.paginator import Paginator
from .filters import (
    CompileFilter,
    # AttendanceDownloadFilter,
    # CompiledAttendanceDownloadFilter,
    AttendanceFilter,
    RawAttendanceFilter,
)
import datetime
from .tasks import save_recompiled, sync_raw_attendance, compile
from .forms import RecompileForm
from django.db.models import Count
from django.contrib.auth.decorators import user_passes_test
from leave.tasks import calculate_total_days


@login_required
def dashboard(request):
    attendances = Attendance.objects.all()
    employees = Employee.objects.all()
    shifts = Shift.objects.all()
    devices = Device.objects.all()
    overtimes = Overtime.objects.all()
    leaves = Leave.objects.all()
    holidays = Holiday.objects.all()
    approvals = (
        overtimes.filter(approved=False, rejected=False).count()
        + leaves.filter(approved=False, rejected=False).count()
        + holidays.filter(approved=False, rejected=False).count()
    )

    start = datetime.datetime.today().date() - datetime.timedelta(days=30)
    most_absents = (
        employees.filter(
            attendances__status="Absent", attendances__check_in_date__gte=start
        )
        .annotate(absent_count=Count("attendances"))
        .order_by("-absent_count")[:20]
    )

    new_employees = Employee.objects.filter(status="Active").order_by("-employee_id")[
        :10
    ]

    return render(
        request,
        "dashboard.html",
        {
            "attendances": attendances,
            "employees": employees,
            "shifts": shifts,
            "devices": devices,
            "overtimes": overtimes,
            "leaves": leaves,
            "holidays": holidays,
            "approvals": approvals,
            "new_employees": new_employees,
            "most_absents": most_absents,
        },
    )


@login_required
def attendance_list(request):
    user = request.user.profile
    if user.role == "HR" or user.role == "ADMIN":
        # if user.device:
        #     attendances = Attendance.objects.filter(
        #         approved=True, deleted=False, device=user.device
        #     ).order_by("-check_in_date", "check_in_time")
        # else:
        attendances = Attendance.objects.filter(approved=True, deleted=False).order_by(
            "-check_in_date", "check_in_time"
        )
    else:
        attendances = Attendance.objects.filter(
            approved=True, deleted=False, employee__department__in=user.manages.all()
        ).order_by("employee")

    attendance_filter = AttendanceFilter(
        request.GET, queryset=attendances, prefix="main"
    )
    attendances = attendance_filter.qs

    download_get_data = request.GET.copy()
    for key in list(download_get_data.keys()):
        if key.startswith("main-"):
            new_key = "download-" + key[len("main-") :]
            download_get_data[new_key] = download_get_data.pop(key)[0]

    download_filter = AttendanceFilter(
        download_get_data, queryset=attendances, prefix="download"
    )
    paginated = Paginator(attendances, 30)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    context = {
        "download_filter": download_filter,
        "filter": attendance_filter,
        "attendances": attendances,
        "page": page,
    }
    return render(
        request,
        "attendance/list.html",
        context,
    )


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def compile_view(request):
    request_device = request.user.profile.device
    if not request_device:
        return redirect("attendance:raw_attendance")
    if DailyRecord.objects.filter(device=request_device).exists():
        daily_records = DailyRecord.objects.filter(device=request_device).latest("date")
        last_date = daily_records.date
        current_date = daily_records.date + datetime.timedelta(days=1)
    else:
        daily_records = []
        last_date = ""
        current_date = datetime.date.today()

    # no_shift = Employee.objects.filter(
    #     shift=None, status="Active", device=request_device
    # ).count()
    attendances = Attendance.objects.filter(
        approved=False, device=request_device
    ).order_by("check_in_time")
    compile_filter = CompileFilter(request.GET, queryset=attendances, prefix="main")
    attendances = compile_filter.qs
    download_get_data = request.GET.copy()
    for key in list(download_get_data.keys()):
        if key.startswith("main-"):
            new_key = "download-" + key[len("main-") :]
            download_get_data[new_key] = download_get_data.pop(key)[0]
    download_filter = CompileFilter(
        download_get_data, queryset=attendances, prefix="download"
    )

    paginated = Paginator(attendances, 30)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    context = {
        "page": page,
        "last_date": last_date,
        "current_date": current_date,
        # "no_shift": no_shift,
        "filter": compile_filter,
        "download_filter": download_filter,
    }
    return render(request, "attendance/compile/list.html", context)


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def compile_attendance(request):
    request_device = request.user.profile.device
    if DailyRecord.objects.filter(device=request_device).exists():
        date = DailyRecord.objects.filter(device=request_device).latest(
            "date"
        ).date + datetime.timedelta(days=1)
    else:
        date = datetime.date.today() - datetime.timedelta(days=1)
    try:
        if date > datetime.date.today():
            pass
        else:
            compile(
                date=date,
                employees=None,
                pattern=None,
                recompiled=False,
                request_device=request_device,
            )
    except Exception as e:
        print(e)
    return redirect("attendance:compile_view")


@login_required
def download_compiled_attendance(request):
    user = request.user.profile
    attendances = Attendance.objects.filter(
        approved=False, deleted=False, recompiled=False, device=user.device
    ).order_by("-check_in_time")
    form = CompileFilter(
        data=request.POST,
        queryset=attendances,
        prefix="download",
    )
    attendances = form.qs
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="compiled.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = [
        "Employee",
        "Device",
        "Current pattern",
        "Check in date",
        "Check out date",
        "Check in time",
        "Check out time",
        "Worked_hours",
        "Check in type",
        "Check out type",
        "Status",
        "Leave type",
        "Half Day",
    ]
    ws.append(headers)

    for attendance in attendances.order_by("-check_in_date"):
        half_day = ""
        device = ""
        leave_type = ""

        if attendance.device:
            device = attendance.device.name
        if attendance.leave_type:
            leave_type = attendance.leave_type.name
            leave = Leave.objects.filter(
                employee=attendance.employee,
                approved=True,
                start_date__lte=attendance.check_in_date,
                end_date__gte=attendance.check_in_date,
            ).first()
            if leave.half_day:
                half_day = True
        else:
            half_day = False
        ws.append(
            [
                attendance.employee.name,
                device,
                attendance.current_pattern.name,
                attendance.check_in_date,
                attendance.check_out_date,
                attendance.check_in_time,
                attendance.check_out_time,
                attendance.worked_hours,
                attendance.check_in_type,
                attendance.check_out_type,
                attendance.status,
                leave_type,
                half_day,
            ]
        )
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def delete_compiled_attendance(request):
    request_device = request.user.profile.device
    attendances = Attendance.objects.filter(
        device=request_device,
        approved=False,
        recompiled=False,
    )
    if attendances:
        for attendance in attendances:
            attendance.delete()
    return redirect("attendance:compile_view")


@login_required
def download_attendance(request):
    user = request.user.profile
    if user.role == "HR" or user.role == "ADMIN":
        attendances = Attendance.objects.filter(approved=True, deleted=False).order_by(
            "-check_in_date", "-check_in_time"
        )
    else:
        attendances = Attendance.objects.filter(
            approved=True, deleted=False, employee__department__in=user.manages.all()
        ).order_by("-check_in_date", "-check_in_time")
    form = AttendanceFilter(data=request.POST, queryset=attendances, prefix="download")
    attendances = form.qs
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="attendance.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = [
        "Employee",
        "Device",
        "Current pattern",
        "Check in date",
        "Check out date",
        "Check in time",
        "Check out time",
        "Worked_hours",
        "Check in type",
        "Check out type",
        "Status",
        "Leave type",
        "Half Day",
    ]
    ws.append(headers)

    for attendance in attendances.order_by("-check_in_date"):
        if attendance.device:
            device = attendance.device.name
        else:
            device = ""
        # if attendance.leave_type:
        #     leave_type = attendance.leave_type.name
        # else:
        leaves = Leave.objects.filter(
            employee=attendance.employee,
            approved=True,
            start_date__lte=attendance.check_in_date,
            end_date__gte=attendance.check_in_date,
        )
        if leaves:
            leave_type = leaves.first().leave_type.name
            half_day = leaves.first().half_day
        else:
            leave_type = ""
            half_day = False

        ws.append(
            [
                attendance.employee.name,
                device,
                attendance.current_pattern.name,
                attendance.check_in_date,
                attendance.check_out_date,
                attendance.check_in_time,
                attendance.check_out_time,
                attendance.worked_hours,
                attendance.check_in_type,
                attendance.check_out_type,
                attendance.status,
                leave_type,
                half_day,
            ]
        )
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def raw_attendance_list(request):
    # attendances = RawAttendance.objects.all().order_by("-date", "-time")[:1000]
    request_device = request.user.profile.device

    base_queryset = RawAttendance.objects.filter(device=request_device)
    attendance_filter = RawAttendanceFilter(
        request.GET, queryset=base_queryset, prefix="main"
    )
    attendances = attendance_filter.qs
    download_get_data = request.GET.copy()
    for key in list(download_get_data.keys()):
        if key.startswith("main-"):
            new_key = "download-" + key[len("main-") :]
            download_get_data[new_key] = download_get_data.pop(key)[0]
    download_filter = RawAttendanceFilter(
        download_get_data, queryset=attendances, prefix="download"
    )
    paginated = Paginator(attendances, 30)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    context = {
        "filter": attendance_filter,
        "download_filter": download_filter,
        "page": page,
    }
    return render(request, "attendance/raw_attendance/list.html", context)


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def get_raw_data(request):
    request_device = request.user.profile.device
    if request_device:
        sync_raw_attendance(request_device=request_device.id)
    else:
        sync_raw_attendance()
    return redirect("attendance:raw_attendance")


@login_required
def download_raw_data(request):
    user = request.user.profile
    if user.role == "HR" or user.role == "ADMIN":
        raw_datas = RawAttendance.objects.all()
    else:
        raw_datas = RawAttendance.objects.filter(
            employee__department__in=user.manages.all()
        )
    form = RawAttendanceFilter(
        data=request.POST,
        queryset=raw_datas,
        prefix="download",
    )
    raw_datas = form.qs
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="Raw data.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw data"

    headers = [
        "Employee",
        "Department",
        "Device",
        "Date",
        "Time",
    ]
    ws.append(headers)

    for raw_data in raw_datas.order_by("-date", "-employee__name"):
        if raw_data.employee.department:
            department_name = raw_data.employee.department.name
        else:
            department_name = None

        ws.append(
            [
                raw_data.employee.name,
                department_name,
                raw_data.device.name,
                raw_data.date,
                raw_data.time,
            ]
        )
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def recompile_view(request):
    request_device = request.user.profile.device
    if not request_device:
        return redirect("attendance:attendances")
    recompile_form = RecompileForm()
    qt = request.POST.getlist("employees")
    print(qt)
    employees = Employee.objects.filter(
        employee_id__in=qt, device=request_device
    ).order_by("name")
    attendances = Attendance.objects.filter(
        recompiled=True, approved=False, deleted=False, device=request_device
    ).order_by("-check_in_date", "-check_in_time")
    paginated = Paginator(attendances, 30)
    page_number = request.GET.get("page")
    page = paginated.get_page(page_number)
    context = {
        "recompile_form": recompile_form,
        "attendances": attendances,
        "page": page,
        "employees": employees,
    }
    return render(request, "attendance/recompile/list.html", context)


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def recompile(request):
    request_device = request.user.profile.device

    employees = Employee.objects.filter(
        id__in=request.POST.getlist("employees"), device=request_device
    ).values_list("employee_id", flat=True)
    recompile_form = RecompileForm(request.POST)
    if recompile_form.is_valid():
        pattern = recompile_form.cleaned_data["pattern"]
        date = recompile_form.cleaned_data["date"]
        if pattern:
            id = pattern.id
        else:
            id = None
        compile(
            date=date,
            employees=employees,
            pattern=id,
            recompiled=True,
            request_device=request_device,
        )
    return redirect("attendance:recompile_view")


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def select_for_recompile(request):
    request_device = request.user.profile.device

    employees = Employee.objects.filter(
        status="Active", shift__isnull=False, device=request_device
    ).order_by("name")
    employee_filter = EmployeeFilter(request.GET, queryset=employees)
    employees = employee_filter.qs

    context = {
        "employees": employees,
        "employee_filter": employee_filter,
    }
    return render(request, "attendance/recompile/select.html", context)


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def save_recompile(request):
    save_recompiled(request)
    return redirect("attendance:recompile_view")


@login_required
@user_passes_test(lambda u: u.profile.role == "HR" or u.profile.role == "ADMIN")
def cancel_recompile(request):
    request_device = request.user.profile.device
    attendances = Attendance.objects.filter(
        recompiled=True, approved=False, deleted=False, device=request_device
    )

    for attendance in attendances:
        try:
            deleted_attendance = Attendance.objects.filter(
                check_in_date=attendance.check_in_date,
                employee=attendance.employee,
                deleted=True,
            )
            deleted_attendance.deleted = False
            deleted_attendance.save()
        except:
            pass
        attendance.delete()
    return redirect("attendance:recompile_view")
